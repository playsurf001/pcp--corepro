"""
Inspeção profunda do legado:
- Fórmulas por aba
- Lista completa de cadastros
- Macros VBA (se conseguirmos extrair)
- Estrutura detalhada da OP/BD da OP
"""
import os
from openpyxl import load_workbook

SRC = "/home/user/webapp/legado.xlsx"

def show_formulas(ws, label, rows=30, cols=20):
    print(f"\n--- Fórmulas em [{label}] (até {rows}x{cols}) ---")
    found = 0
    for r in range(1, min(ws.max_row, rows) + 1):
        for c in range(1, min(ws.max_column, cols) + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                print(f"  {cell.coordinate}: {cell.value[:120]}")
                found += 1
                if found > 40:
                    print("  ... (mais de 40 fórmulas — interrompido)")
                    return

def show_full_range(ws, label, max_row, max_col):
    print(f"\n=== DADOS COMPLETOS [{label}] ===")
    # cabeçalho
    hdr = [ws.cell(row=1, column=c).value for c in range(1, max_col+1)]
    print("HEADER:", hdr)
    for r in range(2, max_row+1):
        row = [ws.cell(row=r, column=c).value for c in range(1, max_col+1)]
        if any(v not in (None, "") for v in row):
            print(f"  R{r}: {row}")

def main():
    wb_f = load_workbook(SRC, data_only=False, keep_vba=True)  # com fórmulas
    wb_v = load_workbook(SRC, data_only=True, keep_vba=True)   # com valores

    # 1) fórmulas de cada aba importante
    abas_formulas = [
        "Seq. Operacional", "Consulta Seq. Operacional", "Transf Seq Op.",
        "BD Seq Op", "Ficha Acompanhamento", "Balanceamento",
        "OP", "Consulta OP", "BD da OP", "Menu"
    ]
    for nome in abas_formulas:
        if nome in wb_f.sheetnames:
            show_formulas(wb_f[nome], nome, rows=30, cols=25)

    # 2) Cadastros completos
    for nome, mr, mc in [
        ("Cliente", 10, 3),
        ("Ref", 10, 3),
        ("Operações", 22, 5),
        ("Máquinas", 15, 5),
        ("Aparelhos", 10, 5),
        ("Cores", 12, 3),
        ("Tamanhos", 25, 3),
    ]:
        if nome in wb_v.sheetnames:
            show_full_range(wb_v[nome], nome, mr, mc)

    # 3) BD Seq Op e BD da OP (para ver o que já foi gravado)
    print("\n" + "="*80)
    print("BD Seq Op — até 20x10")
    ws = wb_v["BD Seq Op"]
    for r in range(1, min(20, ws.max_row)+1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 11)]
        if any(v not in (None, "") for v in row):
            print(f"  R{r}: {row}")

    print("\n" + "="*80)
    print("BD da OP — até 10x30 (todas colunas)")
    ws = wb_v["BD da OP"]
    for r in range(1, min(10, ws.max_row)+1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 31)]
        if any(v not in (None, "") for v in row):
            print(f"  R{r}: {row}")

    # 4) Macros VBA — tentar extrair texto
    print("\n" + "="*80)
    print("ANÁLISE VBA")
    if wb_v.vba_archive:
        print(f"  vba_archive presente: {type(wb_v.vba_archive)}")
        # Tentar listar conteúdos
        import zipfile, io
        try:
            with zipfile.ZipFile(SRC) as z:
                for n in z.namelist():
                    if 'vba' in n.lower() or n.endswith('.bin'):
                        print(f"    arquivo: {n}  size={z.getinfo(n).file_size}")
        except Exception as e:
            print(f"  (zip err {e})")

if __name__ == "__main__":
    main()
