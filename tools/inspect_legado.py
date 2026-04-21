"""
Inspetor da planilha legada: estrutura bruta.
- lista abas
- dimensões
- primeiras linhas (preview)
- detecta merges, validações, tabelas, nomes definidos, macros
"""
import os, json, sys
from openpyxl import load_workbook

SRC = "/home/user/webapp/legado.xlsx"

def preview_cells(ws, max_rows=25, max_cols=20):
    rows = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        row = []
        for c in range(1, min(ws.max_column, max_cols) + 1):
            v = ws.cell(row=r, column=c).value
            row.append(v)
        rows.append(row)
    return rows

def main():
    print(f"Arquivo: {SRC}")
    print(f"Tamanho: {os.path.getsize(SRC)} bytes\n")

    # Carrega com dados calculados visíveis
    wb = load_workbook(SRC, data_only=True, keep_vba=True)
    print(f"Contém macros? {wb.vba_archive is not None}")
    print(f"Abas ({len(wb.sheetnames)}): {wb.sheetnames}\n")

    # Nomes definidos
    print("=== NOMES DEFINIDOS ===")
    try:
        for name in wb.defined_names:
            dn = wb.defined_names[name]
            try:
                dests = list(dn.destinations)
            except Exception:
                dests = []
            print(f"  {name} -> {dests}")
    except Exception as e:
        print(f"(erro listando nomes: {e})")
    print()

    # Detalhes por aba
    for sh in wb.sheetnames:
        ws = wb[sh]
        print("="*80)
        print(f"ABA: {sh}")
        print(f"  Estado: {ws.sheet_state}  Dim: {ws.max_row} x {ws.max_column}")
        # merges
        if ws.merged_cells.ranges:
            print(f"  Merges ({len(ws.merged_cells.ranges)}): {[str(r) for r in list(ws.merged_cells.ranges)[:10]]}{' ...' if len(ws.merged_cells.ranges)>10 else ''}")
        # tabelas estruturadas
        if getattr(ws, 'tables', None):
            for tname, tref in ws.tables.items():
                print(f"  Tabela estruturada: {tname} -> {tref}")
        # data validations
        try:
            dvs = ws.data_validations.dataValidation
            if dvs:
                for dv in dvs[:5]:
                    print(f"  DV: type={dv.type} formula1={dv.formula1} ranges={[str(x) for x in dv.sqref.ranges][:3]}")
                if len(dvs) > 5:
                    print(f"  ...(+{len(dvs)-5} DVs)")
        except Exception:
            pass

        # Preview
        preview = preview_cells(ws, max_rows=20, max_cols=15)
        print("  PREVIEW (até 20x15):")
        for i, row in enumerate(preview, 1):
            # compacta None
            compact = ["" if v is None else (str(v)[:25]) for v in row]
            print(f"    R{i:02d}: {compact}")
        print()

if __name__ == "__main__":
    main()
